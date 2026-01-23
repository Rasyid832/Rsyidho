from flask import Flask, request, render_template_string, send_from_directory
from openpyxl import Workbook, load_workbook
import qrcode, os, json, time, uuid

app = Flask(__name__)

IP = "10.203.208.10"

# =========================
# STYLE
# =========================
STYLE = """
<style>
body{
    background:#0f172a;
    font-family:Arial;
    display:flex;
    justify-content:center;
    align-items:center;
    height:100vh;
    margin:0;
}
.card{
    background:white;
    padding:30px;
    border-radius:12px;
    width:320px;
    text-align:center;
    box-shadow:0 0 20px rgba(0,0,0,0.3);
}
h2{color:#1e293b;}
input{
    width:100%;
    padding:10px;
    margin-top:10px;
    border-radius:8px;
    border:1px solid #ccc;
}
button{
    width:100%;
    padding:10px;
    margin-top:15px;
    background:#2563eb;
    border:none;
    color:white;
    border-radius:8px;
    font-size:16px;
    cursor:pointer;
}
a{
    display:block;
    margin-top:15px;
    color:#2563eb;
    text-decoration:none;
}
img{
    width:200px;
    margin-top:10px;
}
</style>
"""

EXCEL_FILE = "absensi.xlsx"
TOKEN_FILE = "tokens.json"

# =========================
# Setup file
# =========================
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nama", "Kelas", "Waktu"])
    wb.save(EXCEL_FILE)

if not os.path.exists(TOKEN_FILE):
    with open(TOKEN_FILE, "w") as f:
        json.dump({}, f)

if not os.path.exists("static"):
    os.mkdir("static")

# =========================
# Halaman Guru
# =========================
@app.route("/")
def home():
    return STYLE + """
    <div class="card">
        <h2>Guru Panel</h2>
        <a href="/buat_qr"><button>Buatm4 QR Absensi</button></a>
    </div>
    """

# =========================
# Buat QR
# =========================
@app.route("/buat_qr")
def buat_qr():
    token = str(uuid.uuid4())
    expiry = time.time() + 120

    with open(TOKEN_FILE) as f:
        tokens = json.load(f)

    tokens[token] = {"expiry": expiry}

    with open(TOKEN_FILE, "w") as f:
        json.dump(tokens, f)

    link = f"http://{IP}:5000/scan?token={token}"
    img = qrcode.make(link)
    img.save("static/qr.png")

    return STYLE + f"""
    <div class="card">
        <h2>QR Absensi</h2>
        <p>Berlaku 2 menit</p>
        <img src="/qr">
        <p style="font-size:12px">{link}</p>
        <a href="/">Kembali</a>
    </div>
    """

# =========================
# Tampilkan QR
# =========================
@app.route("/qr")
def show_qr():
    return send_from_directory("static", "qr.png")

# =========================
# Murid Scan QR
# =========================
@app.route("/scan")
def scan():
    token = request.args.get("token")
    return render_template_string(STYLE + """
    <div class="card">
        <h2>ABSEN MURID</h2>
        <form method="post" action="/submit">
            <input type="hidden" name="token" value="{{token}}">
            <input name="nama" placeholder="Nama" required>
            <input name="kelas" placeholder="Kelas" required>
            <button>Absen</button>
        </form>
    </div>
    """, token=token)

# =========================
# Simpan ke Excel
# =========================
@app.route("/submit", methods=["POST"])
def submit():
    token = request.form["token"]
    nama = request.form["nama"]
    kelas = request.form["kelas"]

    with open(TOKEN_FILE) as f:
        tokens = json.load(f)

    if token not in tokens:
        return "QR tidak valid"

    if time.time() > tokens[token]["expiry"]:
        return "QR sudah kadaluarsa"

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nama, kelas, time.strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL_FILE)

    return STYLE + """
    <div class="card">
        <h2>Absensi Berhasil</h2>
        <p>Terima kasih</p>
    </div>
    """

# =========================
app.run(host="0.0.0.0", port=5000, debug=True)
